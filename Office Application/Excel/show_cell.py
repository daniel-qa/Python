import openpyxl

# 使用 load_workbook 讀取 test.xlsx
workbook = openpyxl.load_workbook('test.xlsx')

# 取得第一個工作表
sheet = workbook.worksheets[0]

# 顯示 row總數 及 column總數
print('row總數:', sheet.max_row)
print('column總數:', sheet.max_column)

# 顯示 cell 資料
for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row = i, column = j).value)