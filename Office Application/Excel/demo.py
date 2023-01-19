import openpyxl
from openpyxl.styles import Font 
import os

# os.chdir 是 python 切換到電腦指定路徑的方法
os.chdir(r"D:\\")
# 請填寫自己電腦裡Excel檔案的絕對路徑
wb = openpyxl.load_workbook('produceSales.xlsx')
# 請填寫要處理的Excel檔案名稱
sheet = wb.worksheets[0]

# 要更正的品名與其單價
price_updates_dict = {'Garlic': 1.99}

#使用for loop掃描所有A欄品名，如果比對一致，就進行更正與上色
print("Processing...")
for rowNum in range(2, sheet.max_row, 1):
    produceName = sheet.cell(rowNum, 1).value
    if produceName in price_updates_dict:
        sheet.cell(rowNum, 2).value = price_updates_dict[produceName]
        sheet.cell(rowNum, 2).font = Font(color='FF0000')
# 將結果另存新檔
wb.save('produceSales_update.xlsx')
print("Done!")
